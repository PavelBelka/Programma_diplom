import openpyxl as openpyxl
import serial
import serial.tools.list_ports
import time

uart = serial.Serial()


def main():
    print('Программа для управления установкой.')
    connect()
    menu_level_0()


def menu_level_0():
    print('Выберите действия:')
    print('1)Режим подготовки к измерению')
    print('2)Текущая температура')
    print('3)Установка температуры')
    print('4)Текущее расстояние')
    print('5)Запись графика')
    key = input()
    menu_level_1(key)


def menu_level_1(k):
    if k == '1':
        print('Выставьте нулевую точку:')
        uart_write('A', 'A', 'A')
    if k == '2':
        print('Температура')
        while 1:
            uart.flushInput()
            uart.flushOutput()
            uart_write('B', 'B', 'B')
            time.sleep(0.15)
            data = uart.read(3)
            temp = 0
            if data[0] == 86:
                print('Авария:поврежден датчик температуры')
                break
            if data[0] == 87:
                temp = (data[1] << 4) | (data[2] >> 4)
            print('Температура: ' + str(temp))
    if k == '3':
        temp = input('Требуемая температура: ')
        uart_write([67], [int(temp) >> 8], [int(temp) & 0xFF])
    if k == '5':
        print("Создаем файл Excel")
        wb = openpyxl.Workbook()
        print("Создаем лист График")
        sheet = wb.create_sheet('График', 0)
        sheet['A1'] = 'Время'
        sheet['B1'] = 'Температура'
        print("Начинаем измерение...")
        print("Включаем систему")
        uart_write('E', 'E', 'E')
        tim = -150
        for i in range(2, 103):
            uart_write('B', 'B', 'B')
            data = uart.read(3)
            temp = (data[1] << 4) | (data[2] >> 4)
            tim = tim + 150
            cell_time = sheet.cell(row = i, column = 1)
            cell_temp = sheet.cell(row = i, column = 2)
            cell_time.value = tim
            cell_temp.value = temp
            time.sleep(0.15)
        print("Выключаем систему")
        uart_write('F', 'F', 'F')
        print("Сохраняем файл")
        wb.save('PID data.xlsx')
        print("Готово")
    menu_level_0()


def uart_write(command, data_1, data_2):
    uart.write(command.encode('ASCII'))
    time.sleep(0.01)
    uart.write(data_1.encode('ASCII'))
    time.sleep(0.01)
    uart.write(data_2.encode('ASCII'))


def connect():
    lists = serial.tools.list_ports.comports(include_links=False)
    print('Требуется подключение к устрйоству. Выберите порт для подключения:')
    for port in lists:
        print(port.device)
    port_name = input()
    uart.port = port_name
    uart.baudrate = 38400
    uart.bytesize = serial.EIGHTBITS
    uart.parity = serial.PARITY_NONE
    uart.stopbits = serial.STOPBITS_ONE
    uart.timeout = 2
    uart.write_timeout = 2
    uart.open()
    print('Соединение...')
    if uart.is_open:
        uart.flushInput()
        uart.flushOutput()
        print('Порт открыт. Отправка запроса устройству:')
        j = 0
        while True:
            print("Попытка: " + str(j))
            try:
                uart_write('G', 'Y', 'B')
                print('G')
                print('Y')
                print('B')
            except Exception as e:
                print('Ошибка подключения: ' + str(e))
            time.sleep(1)
            print('Получение ответа от устройства:')
            if uart.in_waiting > 0:
                data = uart.read(3)
            else:
                data = [0, 0, 0]
            for i in range(0, 3):
                print(data[i])
            if (data[0] == 65) and (data[1] == 86) and (data[2] == 69):
                print('Соединение установлено.')
                break
            else:
                print('Отказ')
                j = j + 1
                if j == 4:
                    print('Подключено неизвестное устройство. Переподключите еще раз или выберите другой порт')
                    break
    else:
        print('Не открыт порт')


main()
